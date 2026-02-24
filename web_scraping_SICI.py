import pandas as pd
from playwright.sync_api import sync_playwright
import time
import re
import unicodedata
import openpyxl

ARQUIVO_EXCEL = 'MFE - Planilha de Inserção e Classificação das Funções.xlsx'
NOME_ABA = "Todas as Funções (Editável)"
ORGAO_ALVO = "GBP"

def normalizar(texto):
    """ Exige 100% de match ignorando apenas maiúsculas e acentos """
    if pd.isna(texto) or not texto: return ""
    texto = str(texto).lower()
    texto = "".join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = re.sub(r'[^a-z0-9]', '', texto)
    return texto.strip()

def raspar_dados_sici():
    dados_extraidos = []
    print(f"[*] Iniciando Raspagem no SICI para o órgão: {ORGAO_ALVO}...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.set_default_timeout(60000)
        try:
            page.goto("https://sici.rio.rj.gov.br/PAG/principal.aspx", wait_until="networkidle")
            
            root_link = page.locator(f"a.NohArvorePaginaPrincipal:has-text('{ORGAO_ALVO}')").first
            root_id = root_link.get_attribute("id")
            
            if page.locator(f"img[alt='Expand {ORGAO_ALVO}']").count() > 0:
                page.locator(f"img[alt='Expand {ORGAO_ALVO}']").first.click()
                time.sleep(2)

            if root_id:
                match = re.search(r'(.*)t(\d+)$', str(root_id))
                div_id = f"{match.group(1)}n{match.group(2)}Nodes"
                container = page.locator(f"#{div_id}")
                
                for k in range(container.locator("img[alt^='Expand'] >> visible=true").count()):
                    try: container.locator("img[alt^='Expand'] >> visible=true").first.click(); time.sleep(1)
                    except: pass

                items = container.locator("a.NohArvorePaginaPrincipal")
                print(f"[*] Localizados {items.count()} itens. Iniciando leitura individual...")

                for i in range(items.count()):
                    it = items.nth(i)
                    if not it.is_visible() or it.inner_text().strip() == "0": continue
                    
                    try:
                        it.click()
                        cargo_sici = ""
                        for _ in range(5):
                            time.sleep(1)
                            cargo_el = page.locator("#ContentPlaceHolder1_lblCargo")
                            if cargo_el.is_visible():
                                cargo_sici = cargo_el.inner_text().strip()
                                break
                        
                        area_bruta = page.locator(".LadoDireitoCorpoPaginaPrincipal .Row").first.inner_text().strip()
                        area_sici = re.split(r'\s+/\s+|\n', area_bruta)[0].strip()
                        
                        dados_extraidos.append({"S_Area": area_sici, "S_Cargo": cargo_sici})
                    except:
                        print(f"    [!] Pulei item {i} por demora no carregamento do site.")
                        continue
        finally:
            browser.close()
    return pd.DataFrame(dados_extraidos)

def sincronizar_excel_local(df_sici):
    print(f"\n[*] Lendo a planilha local '{ARQUIVO_EXCEL}'...")
    try:
        df_excel = pd.read_excel(ARQUIVO_EXCEL, sheet_name=NOME_ABA)
    except FileNotFoundError:
        print(f"[ERRO] O arquivo '{ARQUIVO_EXCEL}' não foi encontrado.")
        return

    col_orgao = df_excel.columns[1]
    df_orgao = df_excel[df_excel[col_orgao] == ORGAO_ALVO]

    linhas_para_excluir = []
    novos_registros = []

    print("[*] Cruzando dados com 100% de precisão (Match Exato)...")

    # 1. Indexando dados do SICI
    sici_keys = {}
    for _, rs in df_sici.iterrows():
        chave_si = normalizar(rs['S_Area'] + rs['S_Cargo'])
        sici_keys[chave_si] = {"area": rs['S_Area'], "cargo": rs['S_Cargo']}

    # 2. Indexando dados da Planilha
    excel_keys = {}
    for idx, row in df_orgao.iterrows():
        area_ex = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
        cargo_ex = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
        
        # Ignora linhas fantasmas/vazias que o Excel as vezes cria
        if not normalizar(area_ex) and not normalizar(cargo_ex):
            continue
            
        chave_ex = normalizar(area_ex + cargo_ex)
        excel_keys[chave_ex] = {"area": area_ex, "cargo": cargo_ex, "linha": idx + 2}

    # 3. Varredura: DELETAR (Está na Planilha mas NÃO está no SICI)
    for chave_ex, dados_ex in excel_keys.items():
        if chave_ex not in sici_keys:
            linhas_para_excluir.append(dados_ex['linha'])
            # NOVA FORMATAÇÃO: Título/Área primeiro
            print(f"    [-] EXCLUSÃO -> Área/Título: '{dados_ex['area']}' | Cargo: '{dados_ex['cargo']}' (Linha {dados_ex['linha']})")

    # 4. Varredura: ADICIONAR (Está no SICI mas NÃO está na Planilha)
    for chave_si, dados_si in sici_keys.items():
        if chave_si not in excel_keys:
            nova_linha = ["", ORGAO_ALVO, "", dados_si['area'], dados_si['cargo'], ""]
            novos_registros.append(nova_linha)
            # NOVA FORMATAÇÃO: Título/Área primeiro
            print(f"    [+] ADIÇÃO -> Área/Título: '{dados_si['area']}' | Cargo: '{dados_si['cargo']}'")

    # 5. APLICAÇÃO FÍSICA NO EXCEL
    print(f"\n[*] Aplicando alterações no arquivo '{ARQUIVO_EXCEL}'...")
    try:
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        ws = wb[NOME_ABA]
        
        # Deleta as linhas de baixo para cima
        if linhas_para_excluir:
            linhas_para_excluir.sort(reverse=True)
            for linha in linhas_para_excluir:
                ws.delete_rows(linha)
                print(f"    [X] Linha {linha} excluída fisicamente.")
        else:
            print("    [✓] Nenhum cargo pendente de exclusão.")

        # Adiciona novos registros ignorando os zeros da Coluna A
        if novos_registros:
            ultima_linha = ws.max_row
            while ultima_linha > 1 and ws.cell(row=ultima_linha, column=2).value is None:
                ultima_linha -= 1
            
            linha_alvo = ultima_linha + 1

            for registro in novos_registros:
                for col_idx, valor in enumerate(registro, start=1):
                    if valor: 
                        ws.cell(row=linha_alvo, column=col_idx, value=valor)
                linha_alvo += 1
            print(f"    [✓] {len(novos_registros)} novos cargos injetados na posição correta.")
        else:
            print("    [✓] Nenhum cargo novo encontrado para adicionar.")

        wb.save(ARQUIVO_EXCEL)
        print("\n[!] Operação concluída e arquivo salvo com sucesso!")

    except PermissionError:
        print(f"\n[ERRO FATAL] O arquivo '{ARQUIVO_EXCEL}' está aberto. Feche-o e tente novamente.")
    except Exception as e:
        print(f"\n[ERRO] Falha no openpyxl: {e}")

if __name__ == "__main__":
    portal_data = raspar_dados_sici()
    
    if not portal_data.empty:
        sincronizar_excel_local(portal_data)
    else:
        print("[ALERTA] A raspagem falhou ou não encontrou dados.")