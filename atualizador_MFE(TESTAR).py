import pandas as pd
import openpyxl
import unicodedata
import re
import datetime
import os
import tkinter as tk
from tkinter import filedialog

# ---------------- CONFIGURAÇÕES ----------------
ARQUIVO_MFE = 'MFE - Planilha de Inserção e Classificação das Funções.xlsx'
NOME_ABA = "Todas as Funções (Editável)"

# ---------------- UTILITÁRIOS ----------------
def normalizar(texto):
    """ Exige 100% de match ignorando apenas maiúsculas e acentos """
    if pd.isna(texto) or not texto: 
        return ""
    texto = str(texto).lower()
    texto = "".join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = re.sub(r'[^a-z0-9]', '', texto)
    return texto.strip()

def selecionar_arquivo_interface():
    """Abre uma janela para o usuário escolher o arquivo SICI gerado"""
    root = tk.Tk()
    root.withdraw() # Oculta a janela principal do Tkinter
    root.attributes("-topmost", True) # Mantém a janela de seleção na frente
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha extraída do SICI",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    return caminho_arquivo

# ---------------- LÓGICA PRINCIPAL ----------------
def atualizar_planilha_mfe(dados_sici):
    """
    Função principal que pode receber tanto um caminho de arquivo (string)
    quanto um DataFrame do Pandas (quando invocada por outro script).
    """
    # 1. Carregar os dados do SICI
    if isinstance(dados_sici, str):
        print(f"[*] Carregando dados do arquivo SICI: {os.path.basename(dados_sici)}")
        df_sici = pd.read_excel(dados_sici)
    else:
        print("[*] Recebendo dados do SICI diretamente do Web Scraper...")
        df_sici = dados_sici

    if df_sici.empty:
        print("[ALERTA] Os dados do SICI estão vazios. Nenhuma atualização será feita.")
        return

    # 2. Carregar a planilha MFE
    print(f"[*] Lendo a planilha local '{ARQUIVO_MFE}'...")
    try:
        df_mfe = pd.read_excel(ARQUIVO_MFE, sheet_name=NOME_ABA)
    except FileNotFoundError:
        print(f"[ERRO] O arquivo '{ARQUIVO_MFE}' não foi encontrado no diretório atual.")
        return

    # 3. Identificar os órgãos que foram raspados no SICI para filtrar a MFE
    orgaos_presentes_sici = df_sici['órgão'].dropna().unique().tolist()
    orgaos_normalizados_sici = [normalizar(org) for org in orgaos_presentes_sici]

    print(f"[*] Cruzando dados para {len(orgaos_presentes_sici)} órgão(s) com 100% de precisão...")

    # 4. Indexar dados do SICI (Chave Única: Órgão + Área + Cargo)
    sici_keys = {}
    for _, rs in df_sici.iterrows():
        orgao_si = str(rs['órgão']) if pd.notna(rs['órgão']) else ""
        area_si = str(rs['área']) if pd.notna(rs['área']) else ""
        cargo_si = str(rs['cargo']) if pd.notna(rs['cargo']) else ""
        
        chave_si = f"{normalizar(orgao_si)}|{normalizar(area_si)}|{normalizar(cargo_si)}"
        sici_keys[chave_si] = {"orgao": orgao_si, "area": area_si, "cargo": cargo_si}

    # 5. Indexar dados da Planilha MFE (Apenas dos órgãos pertinentes)
    mfe_keys = {}
    for idx, row in df_mfe.iterrows():
        orgao_ex = str(row.iloc[1]) if pd.notna(row.iloc[1]) else "" # Coluna B
        
        # Só processa a linha se pertencer a um órgão que foi raspado agora
        if normalizar(orgao_ex) not in orgaos_normalizados_sici:
            continue

        area_ex = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""  # Coluna D
        cargo_ex = str(row.iloc[4]) if pd.notna(row.iloc[4]) else "" # Coluna E
        
        # Ignora linhas fantasmas/vazias
        if not normalizar(area_ex) and not normalizar(cargo_ex):
            continue
            
        chave_ex = f"{normalizar(orgao_ex)}|{normalizar(area_ex)}|{normalizar(cargo_ex)}"
        mfe_keys[chave_ex] = {"orgao": orgao_ex, "area": area_ex, "cargo": cargo_ex, "linha": idx + 2}

    # 6. Mapear Ações (Exclusões e Adições)
    linhas_para_excluir = []
    dados_removidos = [] # Para o backup
    novos_registros = []

    # Varredura: DELETAR (Está na MFE mas NÃO está no SICI)
    for chave_ex, dados_ex in mfe_keys.items():
        if chave_ex not in sici_keys:
            linhas_para_excluir.append(dados_ex['linha'])
            dados_removidos.append(dados_ex)
            print(f"    [-] EXCLUSÃO -> Órgão: '{dados_ex['orgao']}' | Área: '{dados_ex['area']}' | Cargo: '{dados_ex['cargo']}' (Linha {dados_ex['linha']})")

    # Varredura: ADICIONAR (Está no SICI mas NÃO está na MFE)
    for chave_si, dados_si in sici_keys.items():
        if chave_si not in mfe_keys:
            # Layout esperado: [Vazio, Órgão, Vazio, Área, Cargo, Vazio...]
            nova_linha = ["", dados_si['orgao'], "", dados_si['area'], dados_si['cargo'], ""]
            novos_registros.append(nova_linha)
            print(f"    [+] ADIÇÃO -> Órgão: '{dados_si['orgao']}' | Área: '{dados_si['area']}' | Cargo: '{dados_si['cargo']}'")

    # 7. APLICAÇÃO FÍSICA NO EXCEL MFE
    print(f"\n[*] Aplicando alterações no arquivo '{ARQUIVO_MFE}'...")
    try:
        wb = openpyxl.load_workbook(ARQUIVO_MFE)
        ws = wb[NOME_ABA]
        
        # Deleta as linhas de baixo para cima (evita problemas de deslocamento de índice)
        if linhas_para_excluir:
            linhas_para_excluir.sort(reverse=True)
            for linha in linhas_para_excluir:
                ws.delete_rows(linha)
            print(f"    [✓] {len(linhas_para_excluir)} cargos excluídos fisicamente da MFE.")
        else:
            print("    [✓] Nenhum cargo pendente de exclusão na MFE.")

        # Adiciona novos registros no final das áreas preenchidas
        if novos_registros:
            ultima_linha = ws.max_row
            while ultima_linha > 1 and ws.cell(row=ultima_linha, column=2).value is None:
                ultima_linha -= 1
            
            linha_alvo = ultima_linha + 1

            for registro in novos_registros:
                for col_idx, valor in enumerate(registro, start=1):
                    if valor: 
                        ws.cell(row=linha_alvo, column=col_idx, value=valor)
                linha_alvo += 1
            print(f"    [✓] {len(novos_registros)} novos cargos injetados.")
        else:
            print("    [✓] Nenhum cargo novo para adicionar.")

        wb.save(ARQUIVO_MFE)
        print("\n[!] Planilha MFE atualizada salva com sucesso!")

    except PermissionError:
        print(f"\n[ERRO FATAL] O arquivo '{ARQUIVO_MFE}' está aberto no Excel. Feche-o e tente novamente.")
        return
    except Exception as e:
        print(f"\n[ERRO] Falha ao manipular o arquivo via openpyxl: {e}")
        return

    # 8. SALVAR BACKUP DOS REMOVIDOS
    if dados_removidos:
        agora = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        arquivo_removidos = f"removidos_MFE_{agora}.xlsx"
        df_removidos = pd.DataFrame(dados_removidos)
        # Reorganiza as colunas e remove a informação de linha interna do script
        df_removidos = df_removidos[['orgao', 'area', 'cargo']]
        df_removidos.rename(columns={'orgao': 'Órgão', 'area': 'Área/Título', 'cargo': 'Cargo'}, inplace=True)
        
        df_removidos.to_excel(arquivo_removidos, index=False)
        print(f"[!] Planilha de auditoria '{arquivo_removidos}' gerada com sucesso contendo as exclusões.")

# ---------------- EXECUÇÃO DIRETA ----------------
if __name__ == "__main__":
    arquivo_selecionado = selecionar_arquivo_interface()
    
    if arquivo_selecionado:
        atualizar_planilha_mfe(arquivo_selecionado)
    else:
        print("Operação cancelada. Nenhum arquivo foi selecionado.")