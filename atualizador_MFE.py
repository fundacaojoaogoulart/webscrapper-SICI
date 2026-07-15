import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import unicodedata
import re
import datetime
import os
import sys
import shutil
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

import area_negocio_ml # Módulo de Machine Learning
import calculadora_tercis
import config_manager

# ---------------- CONFIGURAÇÕES ----------------
NOME_ABA = "Todas as Funções (Editável)"
COR_ALERTA = "FFFF00" # Amarelo

# ---------------- DICIONÁRIOS INTERNOS (DE-PARA) ----------------





# ---------------- UTILITÁRIOS ----------------
def normalizar(texto):
    if pd.isna(texto) or not str(texto).strip(): 
        return ""
    texto = str(texto).lower()
    texto = "".join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = re.sub(r'[^a-z0-9]', '', texto)
    return texto.strip()

def normalizar_nome(texto):
    if pd.isna(texto) or not str(texto).strip(): 
        return ""
    texto = str(texto).lower()
    texto = "".join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = " ".join(texto.split()) 
    return texto

def inicializar_tkinter():
    root = tk._default_root
    if root is None:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
    return root

def selecionar_arquivo_sici_interface():
    messagebox.showinfo("Selecione um arquivo", "Selecione a planilha extraída do SICI que contém os dados novos.")
    return filedialog.askopenfilename(
        title="1. Selecione a planilha EXTRAÍDA DO SICI",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )


# ---------------- LÓGICA PRINCIPAL ----------------
def atualizar_planilha_mfe(dados_sici):
    
    root = inicializar_tkinter()
    
    if getattr(sys, 'frozen', False):
        app_path = os.path.dirname(sys.executable)
    else:
        app_path = os.path.dirname(os.path.abspath(__file__))
        
    arquivo_base = os.path.join(app_path, "MFE_Base.xlsx")
    arquivo_saida = os.path.join(app_path, "MFE_Atualizada.xlsx")
    
    if not os.path.exists(arquivo_base):
        try:
            log_path = os.path.join(app_path, "debug.txt")
            with open(log_path, "a", encoding="utf-8") as logf:
                logf.write(f"\\n[{datetime.datetime.now()}] Iniciando extração.\\n")
                logf.write(f"app_path: {app_path}\\n")
                
                caminho_interno = os.path.join(app_path, "_internal", "MFE_Base.xlsx")
                logf.write(f"caminho_interno: {caminho_interno}\\n")
                
                if not os.path.exists(caminho_interno) and hasattr(sys, '_MEIPASS'):
                    caminho_interno = os.path.join(sys._MEIPASS, "MFE_Base.xlsx")
                    logf.write(f"caminho_interno MEIPASS: {caminho_interno}\\n")
                    
                existe_interno = os.path.exists(caminho_interno)
                logf.write(f"Existe interno? {existe_interno}\\n")
                
                if existe_interno:
                    # Copiando em modo binário puro para evitar bloqueios de metadados do Windows
                    with open(caminho_interno, 'rb') as src, open(arquivo_base, 'wb') as dst:
                        dst.write(src.read())
                    logf.write(f"Cópia executada com SUCESSO para: {arquivo_base}\\n")
                
                existe_base = os.path.exists(arquivo_base)
                logf.write(f"Existe base após cópia? {existe_base}\\n")
                
        except Exception as e:
            with open(os.path.join(app_path, "debug.txt"), "a", encoding="utf-8") as logf:
                logf.write(f"EXCECAO FATAL NA COPIA: {str(e)}\\n")
            messagebox.showerror("Erro de Extração", f"Falha ao extrair a planilha base:\\nErro: {str(e)}")

    if not os.path.exists(arquivo_base):
        print(f"[ALERTA] Planilha base '{arquivo_base}' não encontrada.")
        messagebox.showwarning("Aviso", f"Planilha base '{arquivo_base}' não encontrada no diretório atual. Operação cancelada.")
        return 

    # --- 1. CARREGAR OS DADOS DO SICI ---
    if isinstance(dados_sici, str):
        print(f"[*] Carregando dados do arquivo SICI: {os.path.basename(dados_sici)}")
        df_sici = pd.read_excel(dados_sici)
    else:
        print("[*] Recebendo dados do SICI diretamente do fluxo principal...")
        df_sici = dados_sici

    if df_sici.empty:
        print("[ALERTA] Os dados do SICI estão vazios.")
        return

    # --- 2. CARREGAR AS CONFIGURAÇÕES GERAIS ---
    config_data = config_manager.ler_config()
    cargos_excecao = config_data['CARGO_EXATO']
    areas_excecao_contem = config_data['AREA_CONTEM']
    dic_tipos = config_data['TIPOS_CARGO']
    dic_macroareas = config_data['MACRO_AREAS']
    
    verificar_ordenadores = messagebox.askyesno(
        "Ordenadores de Despesa", 
        "Deseja verificar por ordenadores de despesa? (Será necessário fornecer a base de dados)"
    )
    
    set_ordenadores = set()
    if verificar_ordenadores:
        messagebox.showinfo("Selecione um arquivo", "Selecione o arquivo que contém a lista de Ordenadores.")
        planilha_ordenadores = filedialog.askopenfilename(
            title="Selecione a base de Ordenadores",
            filetypes=[("Arquivos CSV", "*.csv"), ("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
        )
        if not planilha_ordenadores:
            messagebox.showwarning("Aviso", "Arquivo não selecionado. A verificação será pulada.")
            verificar_ordenadores = False
        else:
            try:
                _, ext = os.path.splitext(planilha_ordenadores)
                
                df_ord = None
                if ext.lower() == '.csv':
                    tentativas = [(';', 'utf-8'), (';', 'latin1'), (',', 'utf-8'), (',', 'latin1')]
                    for sep, enc in tentativas:
                        try:
                            tmp_df = pd.read_csv(planilha_ordenadores, sep=sep, encoding=enc)
                            if len(tmp_df.columns) > 1:
                                df_ord = tmp_df
                                break
                        except:
                            continue
                    if df_ord is None: 
                        df_ord = pd.read_csv(planilha_ordenadores, sep=';', on_bad_lines='skip')
                else:
                    df_ord = pd.read_excel(planilha_ordenadores, sheet_name=0)
                
                col_user = next((c for c in df_ord.columns if normalizar(str(c)) == 'usuario'), None)
                if not col_user:
                    col_user = df_ord.columns[2] if len(df_ord.columns) >= 3 else df_ord.columns[0]
                
                for val in df_ord[col_user].dropna():
                    nome_limpo = normalizar_nome(str(val))
                    if len(nome_limpo) > 2:
                        set_ordenadores.add(nome_limpo)
                        
                print(f"[*] Base de Ordenadores carregada com {len(set_ordenadores)} usuários válidos.")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao processar arquivo de ordenadores:\n{e}")
                verificar_ordenadores = False

    calcular_tercis = messagebox.askyesno(
        "Magnitude do Orçamento", 
        "Deseja calcular o Critério de Magnitude do Orçamento (Tercis)? (Será necessário fornecer a base de empenhos)"
    )
    
    df_tercis = None
    if calcular_tercis:
        messagebox.showinfo("Selecione um arquivo", "Selecione a planilha de Valor Empenhado.")
        planilha_empenhos = filedialog.askopenfilename(
            title="Selecione a base de Empenhos",
            filetypes=[("Arquivos Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos os Arquivos", "*.*")]
        )
        if not planilha_empenhos:
            messagebox.showwarning("Aviso", "Arquivo não selecionado. O cálculo será pulado.")
            calcular_tercis = False
        else:
            try:
                df_tercis = calculadora_tercis.gerar_dataframe_empenhos(planilha_empenhos)
                print("[*] Base de Empenhos processada com sucesso.")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao processar arquivo de empenhos:\n{e}")
                calcular_tercis = False

    # --- 3. MACHINE LEARNING (Predição em Lote) ---
    print("[*] Acionando a Inteligência Artificial para as Áreas de Negócio...")
    areas_unicas_originais = df_sici['área'].fillna("").unique().tolist()
    areas_negocio_previstas = area_negocio_ml.prever_area_negocio(areas_unicas_originais)
    
    # Cria um de-para do modelo: { "Área original": "Área Prevista ML" }
    dic_ml_areas = dict(zip(areas_unicas_originais, areas_negocio_previstas))

    # --- 4. AVALIAÇÃO E INDEXAÇÃO SICI ---
    sici_keys = {}
    lixos = {'vago', 'vaga', 'nao informado', 'sem titular', '-'}
    
    dic_tercis = {}
    if calcular_tercis and df_tercis is not None:
        dic_tercis = dict(zip(df_tercis['Assinatura_Norm'], df_tercis['Texto_Final_Tercis']))
        
    for _, rs in df_sici.iterrows():
        org_original = str(rs.get('órgão', ''))
        org = normalizar(org_original)
        
        esc = normalizar(str(rs.get('escalão', '')))
        
        area_original = str(rs.get('área', ''))
        area = normalizar(area_original)
        
        cargo = normalizar(str(rs.get('cargo', '')))
        
        titular_original = str(rs.get('titular', '')).strip()
        titular_norm = normalizar_nome(titular_original)
        
        chave_si = f"{org}|{esc}|{area}|{cargo}|{titular_norm}"
        
        is_ordenador = False
        if verificar_ordenadores and titular_norm and titular_norm not in lixos and len(titular_norm) > 2:
            is_ordenador = titular_norm in set_ordenadores
            
        texto_ordenador = "2 - Sim" if is_ordenador else "1 - Não"
        
        is_excecao_poder = False
        if cargo in cargos_excecao:
            is_excecao_poder = True
        else:
            for area_excecao in areas_excecao_contem:
                if area_excecao in area: 
                    is_excecao_poder = True
                    break

        if is_ordenador or is_excecao_poder:
            texto_poder_decisao = "2 - Possui poder de decisão sobre alocação de recursos orçamentários no órgão"
        else:
            texto_poder_decisao = "1 - Não possui poder de decisão sobre alocação de recursos orçamentários no órgão"

        tipo_cargo_classificado = dic_tipos.get(cargo, "")
        macro_area_classificada = dic_macroareas.get(org, "")
        
        # Resgata a área classificada pela IA no Dicionário
        area_negocio_classificada = dic_ml_areas.get(area_original, "")

        texto_tercis = ""
        if calcular_tercis:
            if titular_norm and titular_norm not in lixos:
                texto_tercis = dic_tercis.get(titular_norm, "1 - Não ordena despesa | 0 empenho(s)")
            else:
                texto_tercis = "1 - Não ordena despesa | 0 empenho(s)"

        sici_keys[chave_si] = {
            "orgao": org_original,
            "escalao": str(rs.get('escalão', '')),
            "area": area_original,
            "cargo": str(rs.get('cargo', '')),
            "titular": titular_original,
            "texto_ordenador": texto_ordenador,
            "texto_poder_decisao": texto_poder_decisao,
            "tipo_cargo": tipo_cargo_classificado,
            "macro_area": macro_area_classificada,
            "area_negocio": area_negocio_classificada,
            "texto_tercis": texto_tercis
        }

    # --- 5. PREPARAR DADOS PARA A PLANILHA NOVA ---
    novos_registros = [] 

    for chave_si, dados_si in sici_keys.items():
        nova_linha = [""] * 13 
        nova_linha[1] = dados_si['orgao']  
        nova_linha[2] = dados_si['escalao'] 
        nova_linha[3] = dados_si['area']    
        nova_linha[4] = dados_si['cargo']   
        nova_linha[5] = dados_si['tipo_cargo']            # Coluna F 
        nova_linha[6] = dados_si['area_negocio']         # Coluna G (ML)
        nova_linha[7] = dados_si['macro_area']            # Coluna H 
        if dados_si['tipo_cargo']:
            nova_linha[8] = "1 - Não" if dados_si['tipo_cargo'] == "Ouvidor(a)" else "2 - Sim" # Coluna I
        nova_linha[9] = dados_si['texto_ordenador']       # Coluna J
        nova_linha[10] = dados_si.get('texto_tercis', '')  # Coluna K
        nova_linha[11] = dados_si['texto_poder_decisao']  # Coluna L
        nova_linha[12] = dados_si['titular']              # Coluna M 
        
        novos_registros.append(nova_linha)

    # --- 6. ATUALIZAÇÃO FÍSICA NO EXCEL ---
    import shutil
    print(f"\n[*] Criando '{arquivo_saida}' a partir de '{arquivo_base}'...")
    try:
        shutil.copy(arquivo_base, arquivo_saida)
        
        wb = openpyxl.load_workbook(arquivo_saida)
        ws = wb[NOME_ABA]
        
        cel_cabecalho_13 = ws.cell(row=1, column=13)
        if not cel_cabecalho_13.value:
            cel_cabecalho_13.value = "Titular"
        
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        
        if novos_registros:
            linha_alvo = 2
            for registro in novos_registros:
                for col_idx, valor in enumerate(registro, start=1):
                    if valor: 
                        cel_add = ws.cell(row=linha_alvo, column=col_idx, value=valor)
                        # Pinta a Coluna 7 (G) de Amarelo se o ML der alerta em novos registros
                        if col_idx == 7 and str(valor).startswith("⚠️"):
                            cel_add.fill = PatternFill(start_color=COR_ALERTA, end_color=COR_ALERTA, fill_type="solid")
                        elif col_idx == 9 and str(valor) == "1 - Não":
                            cel_add.fill = PatternFill(start_color=COR_ALERTA, end_color=COR_ALERTA, fill_type="solid")
                            
                linha_alvo += 1

        wb.save(arquivo_saida)
        print("\n[!] Planilha MFE Atualizada gerada com sucesso!")
        
    except PermissionError:
        messagebox.showerror("Erro: Arquivo aberto", f"O arquivo {arquivo_saida} ou {arquivo_base} está aberto no Excel, feche-o e tente novamente.")
        return
    except Exception as e:
        messagebox.showerror("Erro: openpyxl", f"Falha ao manipular o arquivo via openpyxl:\n {e}")
        return

    mensagem_resumo = f"Processo finalizado com sucesso!\n\n✅ Nova planilha '{arquivo_saida}' gerada com {len(novos_registros)} registros."
    mensagem_resumo += f"\n\n⚙️ A coluna G (Área de Negócio) foi preenchida utilizando Inteligência Artificial."
    
    if verificar_ordenadores:
        mensagem_resumo += f"\n⚙️ As colunas J e L foram validadas conforme a capacidade de alocar/gerir despesas para cada titular."
    if calcular_tercis:
        mensagem_resumo += f"\n⚙️ A coluna K foi preenchida com a Magnitude do Orçamento (Tercis)."

    messagebox.showinfo("Atualização Concluída", mensagem_resumo)

if __name__ == "__main__":
    root = inicializar_tkinter()
    arquivo_selecionado = selecionar_arquivo_sici_interface()
    
    if arquivo_selecionado:
        atualizar_planilha_mfe(arquivo_selecionado)
